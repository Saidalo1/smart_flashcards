"""
Решение системы линейных уравнений с матрицей 4x4
Используя функции MINVERSE (МОБР) и MMULT (МУМНОЖ)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Matrix_Solver"

# Матрица A (4x4)
matrix_A = [
    [1, 2, -1, 1],
    [2, 1, 1, 1],
    [1, -1, 2, 1],
    [1, 1, -1, 3]
]

# Вектор b
vector_b = [7, 5, 1, 12]

# Заполняем матрицу A (A1:D4)
for i, row in enumerate(matrix_A, 1):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)

# Заполняем вектор b (F1:F4)
for i, val in enumerate(vector_b, 1):
    ws.cell(row=i, column=6, value=val)

# Обратная матрица A^(-1) в строках 6-9 (A6:D9)
# Формула массива для обратной матрицы
ws["A6"] = "=MINVERSE(A1:D4)"

# Решение x = A^(-1) * b
# Результат в F6:F9
ws["F6"] = "=MMULT(A6:D9,F1:F4)"

# Сохраняем
path = "Zadacha_6_Matrix.xlsx"
wb.save(path)
print(f"Файл создан: {path}")
print("\nВАЖНО: После открытия файла в Excel:")
print("1. Выдели A6:D9, нажми F2, затем Ctrl+Shift+Enter (для обратной матрицы)")
print("2. Выдели F6:F9, нажми F2, затем Ctrl+Shift+Enter (для решения)")
